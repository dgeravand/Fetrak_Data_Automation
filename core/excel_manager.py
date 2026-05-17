# ------------------------------------------------------------------------------
# EXCEL MANAGER
# ------------------------------------------------------------------------------
# Handles Excel file operations: create new files, append data, and auto-size columns.
# ------------------------------------------------------------------------------
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class ExcelManager:

    # ------------------------------------------------------------------------------
    # SAFE SHEET
    # ------------------------------------------------------------------------------
    def _safe_sheet(self, name):
        if not name:
            return "Sheet1"
        return name[:31]

    # ------------------------------------------------------------------------------
    # AUTO WIDTH
    # ------------------------------------------------------------------------------
    def _auto_width(self, worksheet, df):
        for i, col in enumerate(df.columns):
            # Calculate max length for each column
            max_len = max(
                df[col].astype(str).map(len).max(),
                len(str(col))
            )
            width = max_len + 2

            # In openpyxl, work with worksheet as an object
            column_letter = get_column_letter(i + 1)
            worksheet.column_dimensions[column_letter].width = width

    # ------------------------------------------------------------------------------
    # CREATE
    # ------------------------------------------------------------------------------
    def create_excel(self, df, sheet_name="Sheet1"):
        sheet = self._safe_sheet(sheet_name)
        output = BytesIO()

        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet)
            worksheet = writer.sheets[sheet]

            # For xlsxwriter, use set_column to set column widths
            for i, col in enumerate(df.columns):
                max_len = max(df[col].astype(str).map(len).max(), len(str(col)))
                worksheet.set_column(i, i, max_len + 2)

        output.seek(0)
        return output.getvalue()

    # ------------------------------------------------------------------------------
    # APPEND
    # ------------------------------------------------------------------------------
    def append_excel(self, existing_bytes, df, sheet_name="Sheet1"):
        sheet = self._safe_sheet(sheet_name)

        # 1. Read the existing file with openpyxl to find the starting position
        input_stream = BytesIO(existing_bytes)
        book = load_workbook(input_stream)

        start_row = 0
        header = True
        if sheet in book.sheetnames:
            start_row = book[sheet].max_row
            header = False

        # Close the initial book since it will be reopened in writer
        book.close()

        # 2. Use pd.ExcelWriter in append mode
        # Key point: use input_stream as both destination and source in mode 'a'
        output = BytesIO()
        output.write(existing_bytes)
        output.seek(0)

        with pd.ExcelWriter(output, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            df.to_excel(
                writer,
                sheet_name=sheet,
                startrow=start_row,
                index=False,
                header=header
            )

            # 3. Apply auto_width
            ws = writer.sheets[sheet]
            self._auto_width(ws, df)

        output.seek(0)
        return output.getvalue()
